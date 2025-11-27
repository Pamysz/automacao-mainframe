# Manipulação de arquivos Excel
from openpyxl import load_workbook
# Tarefas do pc
import pyautogui
# Tarefas em navegadores
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import os
# Verifica e executa arquivos
import platform 
import time


# 1. Leitura de arquivo .xlsx

arquivo = load_workbook("C:\Arquivos_Teste\Template_Planilha_Teste.xlsx") 
aba_produtos = arquivo["Planilha1"]

def pegar_produtos(aba_produtos):
    produtos = []
    ultima_linha = aba_produtos.max_row

    for linha in range(2, ultima_linha+1):
        nome_produto = aba_produtos.cell(row=linha, column=1).value

        produto = {
            'Produto': nome_produto,
            'Valor': aba_produtos.cell(row=linha, column=2).value,
            'Descricao': aba_produtos.cell(row=linha, column=3).value
        }
        if not produto:
            break


        produtos.append(produto)

    return produtos


lista_produtos = pegar_produtos(aba_produtos)


# 2. Pesquisar produtos na Amazon


# Verificacao de botao de "continuar comprando"
def detectar_clicar(driver, locator, valor_locator, timeout=2):
    try:
        botao = WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((locator, valor_locator))
        )

        botao = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((locator, valor_locator))
        )

        botao.click()
        return True
    except TimeoutException:
        print("Nao foi possivel encontar o botao")
        return False



def pesquisa_amazon(lista_produtos):
     
    driver = webdriver.Edge()


    try:
        resultados = []
        for produto in lista_produtos:
            nome_produto = produto['Produto']

            driver.get("https://www.amazon.com.br/")
            detectar_clicar(driver, By.XPATH, "/html/body/div/div[1]/div[3]/div/div/form/div/div/span/span")
            time.sleep(2)
            caixa_pesquisa = driver.find_element(By.ID, 'twotabsearchtextbox')

            caixa_pesquisa.clear()
            caixa_pesquisa.send_keys(nome_produto)
            caixa_pesquisa.send_keys(Keys.RETURN)
            time.sleep(2)

            try:
                descricao = driver.find_element(By.CSS_SELECTOR, 'div[data-component-type="s-search-result]:not(.AdHolder) a h2 span').text
                time.sleep(1)
            except:
                descricao = "produto nao encontrado"


            try:
                valor = driver.find_element(By.CSS_SELECTOR, 'div[data-component-type="s-search-result]:not(.AdHolder) .a-price .a-offscreen').get_attribute('textContent')
                time.sleep(1)
            except:
                valor = "valor nao encontrado"


            try:
                limite = 35
                if len(descricao) <=limite:
                    descricao_cortada = descricao
                else:
                    descricao_cortada = descricao[:limite]
                    ultimo_espaco = descricao_cortada.rfind(' ')
                    if ultimo_espaco != -1:
                        descricao_cortada = descricao_cortada[:ultimo_espaco]

            except:  
                print(f"Erro no limite")  



            resultados.append({
                'nomeProduto': nome_produto.replace(" ", "")[:10],
                'valor': valor,
                'descricao': descricao_cortada
            })

        return resultados, driver

    except Exception as e:
        driver.quit()
        raise e
    
time.sleep(10)


if lista_produtos:
    resultados, driver = pesquisa_amazon(lista_produtos)
    for resultado in resultados:
        print(f"Produto: {resultado['nomeProduto']}")
        print(f"Valor: {resultado['valor']}")
        print(f"Descricao: {resultado['descricao']}")




# 3. Preenchimento de dados em um sistema Mainframe
def mainframe( nome_mainframe):

    caminho_arquivo = os.path.join('C:/BPATech/tn5250j-0.7.6', nome_mainframe)

    if not os.path.isfile(caminho_arquivo):
        print(f"Arquivo  não existe: {caminho_arquivo} ")
        return False
    
    try:

        if platform.system() == 'Windows':
            os.startfile(caminho_arquivo)


    except Exception as e:
        print(f"Erro{e}")
        return False
    



    time.sleep(10)
    
    try:
        pyautogui.write('BPAPS11')
        pyautogui.press('TAB')
        pyautogui.write('bp4ps25')
        pyautogui.press('Enter')
        time.sleep(2)
        pyautogui.write('strpdm')
        pyautogui.press('Enter')
        time.sleep(2)
        pyautogui.write('3')
        pyautogui.press('Enter')
        time.sleep(2)
        pyautogui.write('QCPPSRC')
        pyautogui.press('Enter')
        pyautogui.press('Enter')

    except Exception as e:
        print(f"Não foi possivel acessar o mainframe{e}")
        return False
    
    # 3.1 Cricacao

def criar(resultados):
    if not resultados:
        print("Nenhum produto para ser criado")
        return
        

    print(f"Iniciando criacao de: {len(resultados)} itens")
    try:
            for i, resultado in enumerate(resultados, 1):
             time.sleep(2)
             pyautogui.press('F6')
             time.sleep(2)
             pyautogui.write(str(resultado.get('nomeProduto', 'SemNome')))
             time.sleep(2)
             pyautogui.press('TAB')
             time.sleep(2)
             pyautogui.write('TXT  ')
             time.sleep(2)
             pyautogui.press('TAB')
             time.sleep(2)
             pyautogui.write(str(resultado.get('valor', '0')))
             time.sleep(2)
             pyautogui.press('Enter')
             time.sleep(2)
             pyautogui.write(str(resultado.get('descricao', '')))
             time.sleep(2)
             pyautogui.press('Enter')
             pyautogui.press('F3')
             pyautogui.press('Enter')

            print(f" {i} itens foram criados com sucesso")

    except Exception as e:
            print(f"Erro na criacao {e}")


mainframe("tn5250j.jar")
criar(resultados)


# 4. Transferencia de dados para Planilha (dados da amazon)

def transferencia(resultados):
    try:
        arquivo = load_workbook("C:\Arquivos_Teste\Template_Planilha_Teste.xlsx") 
        aba_produtos = arquivo["Planilha1"]

        ultima_linha = aba_produtos.max_row

        if len(resultados) != (ultima_linha - 1):
            print("Numero de resultados não corresponde com a planilha")

        for i, resultado in enumerate(resultados, start=2):
            if i > ultima_linha + 1:
               break
            
            try:
                valor_transferido = resultado.get('valor', '')
                descricao_transferida = resultado.get('descricao', '')

                if valor_transferido and descricao_transferida != "produto nao encontrado":
                
                    aba_produtos.cell(row=i, column=2).value = valor_transferido
                    aba_produtos.cell(row=i, column=3).value = descricao_transferida
                    aba_produtos.cell(row=i, column=4).value = "Sucesso"
                    aba_produtos.cell(row=i, column=5).value = "-"
                else:

                    aba_produtos.cell(row=i, column=4).value = "Falha"

                    motivo = []
                    if not valor_transferido or valor_transferido == "valor nao encontrado":
                        motivo.append("Valor nao encontrado")
                    if not descricao_transferida or descricao_transferida == "produto nao encontrado":
                        motivo.append("Descricao nao encontrada")


                    aba_produtos.cell(row=i, column=5).value = ", ".join(motivo) if motivo else "Erro desconhecido"

            except Exception as e:
                aba_produtos.cell(row=i, column=4).value = "Falha"
                aba_produtos.cell(row=i, column=5).value = f"Erro durante a transferencia: {str(e)}"
                print(f"Erro ao transferir dados da linha {i}: {str(e)}")


        arquivo.save("C:\Arquivos_Teste\Template_Planilha_Teste.xlsx")
        print("Dados transferidos com sucesso")


    except Exception as e:
        print(f"Erro na transferencia  {str(e)}")
        for i in range(2, ultima_linha + 1):
            aba_produtos.cell(row=i, column=4).value = "Falha"
            aba_produtos.cell(row=i, column=5).value = f"Erro: {str(e)}"
        arquivo.save("C:\Arquivos_Teste\Template_Planilha_Teste.xlsx")


transferencia(resultados)